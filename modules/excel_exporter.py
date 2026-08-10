import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any

try:
    from modules.tarifas_manager import obtener_tarifa_empleado
except ImportError:
    def obtener_tarifa_empleado(ci, tipo):
        return 0.0


class ExcelExporter:

    @staticmethod
    def _aplicar_autoajuste(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    @staticmethod
    def _safe_float(val, default=0.0):
        try:
            if val is None:
                return default
            if isinstance(val, (int, float)):
                return float(val)
            if hasattr(val, 'year') or hasattr(val, 'hour'):
                return default
            return float(str(val).replace(',', '.'))
        except Exception:
            return default

    @staticmethod
    def exportar_preplanilla_oficial(
        datos_asistencia: List[Dict[str, Any]],
        maestro_empleados: List[Dict[str, Any]],
        periodo: str,
        nombre_archivo: str = "PrePlanilla_Oficial_Fridolin.xlsx"
    ) -> str:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        dict_maestro = {}
        for emp in maestro_empleados:
            ci = str(emp.get('Carnet_Identidad', emp.get('CI', emp.get('Id', '')))).strip()
            if ci:
                dict_maestro[ci] = {
                    'Nombre_Completo': emp.get('Nombre_Completo', emp.get('Nombre', 'SIN NOMBRE')),
                    'Area_Sector': emp.get('Area_Departamento', emp.get('Area_Sector', emp.get('Sector', 'FABRICA'))),
                    'Cargo': emp.get('Rol', emp.get('Cargo', 'OPERARIO')),
                    'Centro_Costo': emp.get('Centro_Costo', 'FRIDOLIN'),
                    'Tipo_Personal': emp.get('Tipo_Personal', 'Fijo')
                }

        font_titulo = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        font_header_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_header_black = Font(name="Calibri", size=11, bold=True, color="000000")
        font_body = Font(name="Calibri", size=10)

        fill_azul = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_gris = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        # ==========================================
        # PESTAÑA 1: FIJOS Y EVENTUALES
        # ==========================================
        ws1 = wb.create_sheet(title="Fijos y Eventuales")
        ws1.views.sheetView[0].showGridLines = True

        ws1.merge_cells("A1:N1")
        ws1["A1"] = f"EMPRESA FRIDOLIN - PRE-PLANILLA DE ASISTENCIA ({periodo})"
        ws1["A1"].font = font_titulo
        ws1["A1"].fill = fill_azul
        ws1["A1"].alignment = align_center

        headers_p1 = [
            ("Nombre_Completo", fill_azul, font_header_white),
            ("Carnet_Identidad", fill_azul, font_header_white),
            ("Area_Sector", fill_azul, font_header_white),
            ("Cargo", fill_azul, font_header_white),
            ("Centro_Costo", fill_azul, font_header_white),
            ("DIAS_TRABAJADOS", fill_verde, font_header_black),
            ("HORAS_EXTRA", fill_verde, font_header_black),
            ("1/2 TURNOS", fill_verde, font_header_black),
            ("REC. NOCTURNO (hrs)", fill_verde, font_header_black),
            ("DOMINICALES", fill_verde, font_header_black),
            ("ATRASOS (min)", fill_rojo, font_header_black),
            ("FALTAS JUSTIFICADA", fill_rojo, font_header_black),
            ("FALTAS INJUSTIFICADA", fill_rojo, font_header_black),
            ("OBSERVACIONES", fill_gris, font_header_black)
        ]

        for col_idx, (h_name, fill_style, font_style) in enumerate(headers_p1, start=1):
            c = ws1.cell(row=3, column=col_idx, value=h_name)
            c.fill = fill_style
            c.font = font_style
            c.alignment = align_center
            c.border = border_thin

        grouped_emp = {}
        for reg in datos_asistencia:
            ci = str(reg.get('Carnet_Identidad', reg.get('ID', reg.get('CI', '')))).strip()
            if not ci:
                continue
            if ci not in grouped_emp:
                grouped_emp[ci] = []
            grouped_emp[ci].append(reg)

        row_idx = 4
        for ci, regs in grouped_emp.items():
            info_m = dict_maestro.get(ci, {
                'Nombre_Completo': regs[0].get('Nombre', 'DESCONOCIDO'),
                'Area_Sector': 'FABRICA',
                'Cargo': 'OPERARIO',
                'Centro_Costo': 'FRIDOLIN',
                'Tipo_Personal': 'Fijo'
            })

            if str(info_m.get('Tipo_Personal', '')).strip().upper() == 'JORNALERO':
                continue

            dias_trab = len([r for r in regs if ExcelExporter._safe_float(r.get('Horas Trabajadas', 0)) > 0 or ExcelExporter._safe_float(r.get('Turnos Computados', 0)) > 0])
            if dias_trab == 0:
                dias_trab = len(regs)

            horas_extra = sum(ExcelExporter._safe_float(r.get('Horas Extras', r.get('Horas_Extra', 0))) for r in regs)

            # ½ TURNOS = cantidad de días con Turnos Computados == 1.5
            medio_turnos = sum(
                1 for r in regs
                if ExcelExporter._safe_float(r.get('Turnos Computados', r.get('1/2 Turnos', 0))) >= 1.5
            )

            # REC. NOCTURNO: solo horas de días nocturnos, con tope razonable
            rec_nocturno = 0.0
            for r in regs:
                turno = str(r.get('Turno Dominante', '')).lower()
                if 'nocturno' in turno:
                    h = ExcelExporter._safe_float(r.get('Horas Nocturnas', r.get('Horas Trabajadas', 0)))
                    # Tope por día: máximo 8 hrs de recargo
                    rec_nocturno += min(h, 8.0)

            dominicales = sum(1 for r in regs if r.get('Es Dominical', False) or str(r.get('Día', '')).lower() == 'domingo')
            atrasos_min = sum(int(ExcelExporter._safe_float(r.get('Atraso (Minutos)', r.get('Atrasos', 0)))) for r in regs)
            faltas_just = sum(int(ExcelExporter._safe_float(r.get('Falta Justificada', 0))) for r in regs)
            faltas_injust = sum(int(ExcelExporter._safe_float(r.get('Falta Injustificada', 0))) for r in regs)

            obs_list = []
            for r in regs:
                obs = str(r.get('Observaciones', r.get('Observacion', ''))).strip()
                if obs and obs not in obs_list and obs.lower() not in ['nan', 'none', '']:
                    obs_list.append(obs)
            observaciones = " | ".join(obs_list)[:250]

            vals_1 = [
                info_m['Nombre_Completo'],
                ci,
                info_m['Area_Sector'],
                info_m['Cargo'],
                info_m['Centro_Costo'],
                dias_trab,
                round(horas_extra, 2),
                medio_turnos,
                round(rec_nocturno, 2),
                dominicales,
                atrasos_min,
                faltas_just,
                faltas_injust,
                observaciones
            ]

            for c_idx, val in enumerate(vals_1, start=1):
                cell = ws1.cell(row=row_idx, column=c_idx, value=val)
                cell.font = font_body
                cell.border = border_thin
                if c_idx in [1, 3, 4, 5, 14]:
                    cell.alignment = align_left
                elif c_idx == 2:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_right
                    if isinstance(val, float):
                        cell.number_format = '#,##0.00'

            row_idx += 1

        ExcelExporter._aplicar_autoajuste(ws1)

        # ==========================================
        # PESTAÑA 2: JORNALEROS
        # ==========================================
        ws2 = wb.create_sheet(title="Jornaleros")
        ws2.views.sheetView[0].showGridLines = True

        ws2.merge_cells("A1:K1")
        ws2["A1"] = f"EMPRESA FRIDOLIN - PLANILLA Y MONETIZACIÓN DE JORNALEROS ({periodo})"
        ws2["A1"].font = font_titulo
        ws2["A1"].fill = fill_azul
        ws2["A1"].alignment = align_center

        headers_p2 = [
            ("Nombre_Completo", fill_azul, font_header_white),
            ("Carnet_Identidad", fill_azul, font_header_white),
            ("Area_Sector", fill_azul, font_header_white),
            ("Diurno Normal 8h (Días)", fill_verde, font_header_black),
            ("Monto Diurno Normal (Bs)", fill_verde, font_header_black),
            ("Diurno 1.5 / 12h (Días)", fill_verde, font_header_black),
            ("Monto Diurno 1.5 (Bs)", fill_verde, font_header_black),
            ("Nocturno Normal 8h (Días)", fill_verde, font_header_black),
            ("Monto Nocturno Normal (Bs)", fill_verde, font_header_black),
            ("Nocturno 1.5 / 12h (Días)", fill_verde, font_header_black),
            ("TOTAL A PAGAR (Bs)", fill_verde, font_header_black)
        ]

        for col_idx, (h_name, fill_style, font_style) in enumerate(headers_p2, start=1):
            c = ws2.cell(row=3, column=col_idx, value=h_name)
            c.fill = fill_style
            c.font = font_style
            c.alignment = align_center
            c.border = border_thin

        row_idx_2 = 4
        for ci, regs in grouped_emp.items():
            info_m = dict_maestro.get(ci, {
                'Nombre_Completo': regs[0].get('Nombre', 'DESCONOCIDO'),
                'Area_Sector': 'FABRICA',
                'Tipo_Personal': 'Jornalero'
            })

            if str(info_m.get('Tipo_Personal', '')).strip().upper() != 'JORNALERO':
                continue

            d_norm = sum(1 for r in regs if str(r.get('Turno Dominante', '')).lower() == 'diurno' and ExcelExporter._safe_float(r.get('Horas Trabajadas', 0)) <= 9)
            d_15 = sum(1 for r in regs if str(r.get('Turno Dominante', '')).lower() == 'diurno' and ExcelExporter._safe_float(r.get('Horas Trabajadas', 0)) > 9)
            n_norm = sum(1 for r in regs if 'nocturno' in str(r.get('Turno Dominante', '')).lower() and ExcelExporter._safe_float(r.get('Turnos Computados', 1)) < 1.5)
            n_15 = sum(1 for r in regs if 'nocturno' in str(r.get('Turno Dominante', '')).lower() and ExcelExporter._safe_float(r.get('Turnos Computados', 0)) >= 1.5)

            t_d_norm = ExcelExporter._safe_float(obtener_tarifa_empleado(ci, "diurno_normal_8h"))
            t_d_15 = ExcelExporter._safe_float(obtener_tarifa_empleado(ci, "diurno_1_5_12h"))
            t_n_norm = ExcelExporter._safe_float(obtener_tarifa_empleado(ci, "nocturno_normal_8h"))
            t_n_15 = ExcelExporter._safe_float(obtener_tarifa_empleado(ci, "nocturno_1_5_12h"))

            m_d_norm = d_norm * t_d_norm
            m_d_15 = d_15 * t_d_15
            m_n_norm = n_norm * t_n_norm
            m_n_15 = n_15 * t_n_15
            total_jornal = m_d_norm + m_d_15 + m_n_norm + m_n_15

            vals_2 = [
                info_m['Nombre_Completo'],
                ci,
                info_m['Area_Sector'],
                d_norm, round(m_d_norm, 2),
                d_15, round(m_d_15, 2),
                n_norm, round(m_n_norm, 2),
                n_15,
                round(total_jornal, 2)
            ]

            for c_idx, val in enumerate(vals_2, start=1):
                cell = ws2.cell(row=row_idx_2, column=c_idx, value=val)
                cell.font = font_body
                cell.border = border_thin
                if c_idx in [1, 3]:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_right
                    if isinstance(val, float):
                        cell.number_format = '#,##0.00'

            row_idx_2 += 1

        ExcelExporter._aplicar_autoajuste(ws2)

        # ==========================================
        # PESTAÑA 3: BONOS PRODUCCIÓN
        # ==========================================
        ws3 = wb.create_sheet(title="Bonos Producción")
        ws3.views.sheetView[0].showGridLines = True

        ws3.merge_cells("A1:G1")
        ws3["A1"] = f"EMPRESA FRIDOLIN - BONOS Y PRODUCCIÓN ADICIONAL ({periodo})"
        ws3["A1"].font = font_titulo
        ws3["A1"].fill = fill_azul
        ws3["A1"].alignment = align_center

        headers_p3 = [
            ("Nombre_Completo", fill_azul, font_header_white),
            ("Carnet_Identidad", fill_azul, font_header_white),
            ("Area_Sector", fill_azul, font_header_white),
            ("Produccion_Normal_Dia", fill_verde, font_header_black),
            ("Unidades_Adicionales", fill_verde, font_header_black),
            ("Monto_Bono_Unidad (Bs)", fill_verde, font_header_black),
            ("TOTAL_BONO_PRODUCCION (Bs)", fill_verde, font_header_black)
        ]

        for col_idx, (h_name, fill_style, font_style) in enumerate(headers_p3, start=1):
            c = ws3.cell(row=3, column=col_idx, value=h_name)
            c.fill = fill_style
            c.font = font_style
            c.alignment = align_center
            c.border = border_thin

        row_idx_3 = 4
        for ci, info_m in dict_maestro.items():
            vals_3 = [
                info_m.get('Nombre_Completo', ''),
                ci,
                info_m.get('Area_Sector', ''),
                0, 0, 0.0, 0.0
            ]
            for c_idx, val in enumerate(vals_3, start=1):
                cell = ws3.cell(row=row_idx_3, column=c_idx, value=val)
                cell.font = font_body
                cell.border = border_thin
                if c_idx in [1, 3]:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_right
            row_idx_3 += 1

        ExcelExporter._aplicar_autoajuste(ws3)

        wb.save(nombre_archivo)
        return nombre_archivo
